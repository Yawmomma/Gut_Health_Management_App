"""
AUSNUT 2023 Import Script
==========================
Imports food nutrient profiles from AUSNUT 2023 xlsx file.

Usage:
    1. Place the file in data/ausnut/:
       - AUSNUT 2023 - Food nutrient profiles.xlsx
    2. Run: python database/import_ausnut_foods.py

The script is idempotent - safe to run multiple times.
Requires: openpyxl (pip install openpyxl)
"""

import sys
import os

# Fix Windows console encoding
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Add parent directory to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

from app import app
from database import db
from models.ausnut import AUSNUTFood, AUSNUTNutrient, AUSNUTFoodNutrient


# Column mapping: (column_index_0based, nutrient_name, unit, group, rank)
# Based on AUSNUT 2023 Food nutrient profiles spreadsheet (row 3 = headers)
NUTRIENT_COLUMNS = [
    (4, 'Energy, with dietary fibre', 'kJ', 'Energy', 1),
    (5, 'Energy, without dietary fibre', 'kJ', 'Energy', 2),
    (6, 'Moisture', 'g', 'Macronutrients', 10),
    (7, 'Protein', 'g', 'Macronutrients', 11),
    (8, 'Total fat', 'g', 'Macronutrients', 12),
    (9, 'Available carbohydrate, with sugar alcohols', 'g', 'Macronutrients', 13),
    (10, 'Available carbohydrate, without sugar alcohols', 'g', 'Macronutrients', 14),
    (11, 'Starch', 'g', 'Macronutrients', 15),
    (12, 'Total sugars', 'g', 'Macronutrients', 16),
    (13, 'Added sugars', 'g', 'Macronutrients', 17),
    (14, 'Free sugars', 'g', 'Macronutrients', 18),
    (15, 'Dietary fibre', 'g', 'Macronutrients', 19),
    (16, 'Alcohol', 'g', 'Macronutrients', 20),
    (17, 'Ash', 'g', 'Macronutrients', 21),
    # Minerals
    (18, 'Calcium', 'mg', 'Minerals', 30),
    (19, 'Iodine', 'ug', 'Minerals', 31),
    (20, 'Iron', 'mg', 'Minerals', 32),
    (21, 'Magnesium', 'mg', 'Minerals', 33),
    (22, 'Phosphorus', 'mg', 'Minerals', 34),
    (23, 'Potassium', 'mg', 'Minerals', 35),
    (24, 'Selenium', 'ug', 'Minerals', 36),
    (25, 'Sodium', 'mg', 'Minerals', 37),
    (26, 'Zinc', 'mg', 'Minerals', 38),
    # Vitamins - A
    (27, 'Retinol', 'ug', 'Vitamins', 40),
    (28, 'Beta-carotene', 'ug', 'Vitamins', 41),
    (29, 'Provitamin A (beta-carotene equivalents)', 'ug', 'Vitamins', 42),
    (30, 'Vitamin A (retinol equivalents)', 'ug', 'Vitamins', 43),
    # Vitamins - B
    (31, 'Thiamin (B1)', 'mg', 'Vitamins', 44),
    (32, 'Riboflavin (B2)', 'mg', 'Vitamins', 45),
    (33, 'Niacin (B3)', 'mg', 'Vitamins', 46),
    (34, 'Pyridoxine (B6)', 'mg', 'Vitamins', 47),
    (35, 'Total folates', 'ug', 'Vitamins', 48),
    (36, 'Dietary folate equivalents', 'ug', 'Vitamins', 49),
    (37, 'Folic acid', 'ug', 'Vitamins', 50),
    (38, 'Natural folate', 'ug', 'Vitamins', 51),
    (39, 'Cobalamin (B12)', 'ug', 'Vitamins', 52),
    # Vitamins - C
    (40, 'Vitamin C', 'mg', 'Vitamins', 53),
    # Vitamins - D
    (41, 'Cholecalciferol (D3)', 'ug', 'Vitamins', 54),
    (42, 'Ergocalciferol (D2)', 'ug', 'Vitamins', 55),
    (43, '25-hydroxycholecalciferol (25-OH D3)', 'ug', 'Vitamins', 56),
    (44, '25-hydroxyergocalciferol (25-OH D2)', 'ug', 'Vitamins', 57),
    (45, 'Vitamin D3 equivalents', 'ug', 'Vitamins', 58),
    # Vitamins - E
    (46, 'Alpha-tocopherol', 'mg', 'Vitamins', 59),
    (47, 'Vitamin E (alpha-tocopherol equivalents)', 'mg', 'Vitamins', 60),
    # Fats
    (48, 'Total saturated fat', 'g', 'Fats', 70),
    (49, 'Total monounsaturated fat', 'g', 'Fats', 71),
    (50, 'Linoleic acid', 'mg', 'Fats', 72),
    (51, 'Alpha-linolenic acid', 'mg', 'Fats', 73),
    (52, 'Eicosapentaenoic acid (EPA)', 'mg', 'Fats', 74),
    (53, 'Docosapentaenoic acid (DPA)', 'mg', 'Fats', 75),
    (54, 'Docosahexaenoic acid (DHA)', 'mg', 'Fats', 76),
    (55, 'Total polyunsaturated fat', 'g', 'Fats', 77),
    (56, 'Total long chain omega-3', 'mg', 'Fats', 78),
    (57, 'Total trans fatty acids', 'mg', 'Fats', 79),
    # Other
    (58, 'Caffeine', 'mg', 'Other', 90),
    (59, 'Cholesterol', 'mg', 'Other', 91),
    (60, 'Tryptophan', 'mg', 'Other', 92),
]


def find_ausnut_file():
    """Find the AUSNUT xlsx file in data/ausnut/"""
    ausnut_dir = os.path.join(os.path.dirname(__file__), '..', 'data', 'ausnut')
    ausnut_dir = os.path.abspath(ausnut_dir)

    if not os.path.exists(ausnut_dir):
        print(f"[ERROR] Directory not found: {ausnut_dir}")
        print("Please create the data/ausnut/ folder and place the AUSNUT xlsx file there.")
        return None

    for filename in os.listdir(ausnut_dir):
        if filename.lower().endswith('.xlsx') and 'ausnut' in filename.lower():
            filepath = os.path.join(ausnut_dir, filename)
            size_mb = os.path.getsize(filepath) / (1024 * 1024)
            print(f"[OK] Found AUSNUT file: {filename} ({size_mb:.1f} MB)")
            return filepath

    print("[ERROR] No AUSNUT xlsx file found in data/ausnut/")
    return None


def import_nutrients():
    """Import nutrient definitions from the column mapping"""
    print("\n[...] Importing AUSNUT nutrient definitions...")
    imported = 0

    for col_idx, name, unit, group, rank in NUTRIENT_COLUMNS:
        existing = AUSNUTNutrient.query.filter_by(name=name).first()
        if not existing:
            nutrient = AUSNUTNutrient(
                name=name,
                unit_name=unit,
                nutrient_group=group,
                rank=rank
            )
            db.session.add(nutrient)
            imported += 1

    db.session.commit()
    print(f"[OK] Imported {imported} new nutrient definitions ({len(NUTRIENT_COLUMNS)} total)")


def import_foods(filepath):
    """Import foods and their nutrient values from the xlsx file"""
    print(f"\n[...] Loading {os.path.basename(filepath)}...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Find the data sheet
    sheet_name = 'Food nutrient profiles'
    if sheet_name not in wb.sheetnames:
        # Try first non-Contents sheet
        for name in wb.sheetnames:
            if name.lower() != 'contents':
                sheet_name = name
                break
    ws = wb[sheet_name]
    print(f"[OK] Using sheet: {sheet_name}")

    # Build nutrient lookup: name -> db id
    nutrient_lookup = {n.name: n.id for n in AUSNUTNutrient.query.all()}

    # Build column->nutrient_id mapping
    col_to_nutrient = {}
    for col_idx, name, unit, group, rank in NUTRIENT_COLUMNS:
        if name in nutrient_lookup:
            col_to_nutrient[col_idx] = nutrient_lookup[name]

    imported_foods = 0
    skipped_foods = 0
    imported_nutrients = 0
    batch_size = 100

    rows = list(ws.iter_rows(min_row=4, values_only=True))  # Data starts at row 4
    total_rows = len(rows)
    print(f"[OK] Found {total_rows} food rows to process")

    for i, row in enumerate(rows):
        if not row or len(row) < 5:
            continue

        survey_id = str(row[0]).strip() if row[0] else None
        if not survey_id:
            continue

        # Check if already imported
        existing = AUSNUTFood.query.filter_by(survey_id=survey_id).first()
        if existing:
            skipped_foods += 1
            continue

        food_name = str(row[3]).strip() if row[3] else None
        if not food_name:
            continue

        public_food_key = str(row[1]).strip() if row[1] else None
        derivation = str(row[2]).strip() if row[2] else None

        food = AUSNUTFood(
            survey_id=survey_id,
            public_food_key=public_food_key,
            derivation=derivation,
            food_name=food_name
        )
        db.session.add(food)
        db.session.flush()

        # Import nutrient values
        for col_idx, nutrient_id in col_to_nutrient.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    amount = float(row[col_idx])
                except (ValueError, TypeError):
                    continue

                food_nutrient = AUSNUTFoodNutrient(
                    food_id=food.id,
                    nutrient_id=nutrient_id,
                    amount=amount
                )
                db.session.add(food_nutrient)
                imported_nutrients += 1

        imported_foods += 1

        if (i + 1) % batch_size == 0:
            db.session.commit()
            progress = ((i + 1) / total_rows) * 100
            print(f"  Progress: {i + 1}/{total_rows} ({progress:.1f}%)")

    db.session.commit()
    wb.close()

    print(f"[OK] Imported {imported_foods} new foods (skipped {skipped_foods} existing)")
    print(f"     {imported_nutrients} nutrient values")
    return imported_foods


def print_summary():
    """Print import summary"""
    print("\n" + "=" * 60)
    print("AUSNUT 2023 IMPORT SUMMARY")
    print("=" * 60)
    print(f"Nutrients:      {AUSNUTNutrient.query.count()}")
    print(f"Foods:          {AUSNUTFood.query.count()}")
    print(f"Food Nutrients: {AUSNUTFoodNutrient.query.count()}")
    print("=" * 60)


def main():
    """Main import function"""
    print("=" * 60)
    print("AUSNUT 2023 FOOD NUTRIENT IMPORT")
    print("=" * 60)

    filepath = find_ausnut_file()
    if not filepath:
        return

    with app.app_context():
        import_nutrients()
        import_foods(filepath)
        print_summary()
        print("\n[OK] Import completed successfully!")


if __name__ == '__main__':
    main()
